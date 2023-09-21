#----------------------------------------------------------------------------------
#
#  Codigo para tratamento dos dados baixadados pelo sistema ANA/HIDROWEB
#
#  O codigo trabalha com os daos baixados pelo script download_ANA_WB.py  
#
#  By Reginadlo Ventura de Sá (reginaldo.venturadesa@gmail.com) 
#
#
#----------------------------------------------------------------------------------
# Data: 24/08/2023 - Finalziação do codigo  (falta calculo do tempo de recorrencia)
# Data: 05/09/2023 - colcoacao das datas no nome de arquivo final.  
#
#
#
#-----------------------------------------------------------------------------------
#
#
import pandas as pd
import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import timedelta
from datetime import datetime
import scipy.stats as stats
#
#
#-------------------------------------------------------------------------------------
#
#  LISTA DE ARQUIVOS BAIXADOS POR downlaod_ANA.py 
#
# Lista de arquivos CSV a serem lidos
#
arquivos_csv = glob.glob('./ANA/*.csv')

#-------------------------------------------------------------------------------------
#
# 
# Função para calcular o tempo de retorno
#  falta fazer 
#
#

# def calcular_tempo_retorno(max_chuva):
    # # Calcular o tempo de retorno usando a distribuição de Gumbel (exemplo)
    # c = 0.4502  # Parâmetro de forma da distribuição de Gumbel
    # loc = max_chuva - 0.5772 / c
    # escala = 1.0 / c
    # tempo_retorno = 1 / (1 - (1 - 0.01)**c)  # Tempo de retorno para 1% de probabilidade
    # return tempo_retorno



def calcular_tempo_retorno(df, chuva_maxima):
    # Ordenar o DataFrame pela chuva máxima em ordem decrescente
    df = df.sort_values(by='chuva', ascending=False)
    
    # Encontrar a posição da primeira ocorrência da chuva máxima fornecida
    posicao_ocorrencia = df[df['chuva'] == chuva_maxima].index[0]
    
    # Calcular a probabilidade de excedência (1 - F)
    probabilidade_excedencia = (posicao_ocorrencia + 1) / (len(df) + 1)
    
    # Calcular o tempo de recorrência usando a fórmula T = (1 / P)
    tempo_retorno = 1 / probabilidade_excedencia
    
    return tempo_retorno

#
# diversaos funcoes
#
#


# Função para calcular a média no período
def calcular_media_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    return periodo['chuva'].mean()


# Função para calcular a soma no período
def calcular_soma_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    return periodo['chuva'].sum()

# Função para calcular a contagem no período
def calcular_conta_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    return periodo['chuva'].count()

# Função para calcular a desvio padrao no período
def calcular_desvpad_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    return periodo['chuva'].std()

    
# Função para calcular a variancia no período
def calcular_variancia_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    return periodo['chuva'].var()
    
# Função para calcular a variancia no período
def calcular_maximo_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    return periodo['chuva'].var()

def calcular_percentil_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]

    # Filtrar valores acima de 0.0 antes de calcular os percentis
    valores_acima_de_zero = periodo[periodo['chuva'] > 0.0]
    
    percentis_dinamico = valores_acima_de_zero['chuva'].quantile([0.05, 0.10, 0.20, 0.25, 0.50, 0.75, 0.90, 0.95, 0.99])
    return percentis_dinamico


# Função para calcular a média no período
def calcular_varios_periodo(df, inicio, fim):
    periodo = df[(df['data'] >= inicio) & (df['data'] <= fim)]
    # Calcular os campos adicionais
    sem_chuva = periodo[periodo['chuva'] == 0.0].shape[0]
    com_chuva = periodo[periodo['chuva'] > 0.0].shape[0]
    sem_dados = periodo[periodo['chuva'].isnull()].shape[0]
    soma_da_chuva = periodo['chuva'].sum()
    status=0
    if len(periodo) == 0:
       status=-1
       num_dados_totais = len(periodo)
       num_dias = (fim - inicio).days  # Calcular número de dias
       num_dados_validos=0
       per_semchuva=0
       per_comchuva=0
       per_semdados=0
       per_dados_validos=0 
    else:
       status=0 
       num_dados_totais = (fim - inicio).days ###len(periodo)
       num_dias = (fim - inicio).days  # Calcular número de dias
       num_dados_validos=sem_chuva+com_chuva
       per_semchuva=sem_chuva/num_dados_totais
       per_comchuva=com_chuva/num_dados_totais
       per_semdados=sem_dados/num_dados_totais
       per_dados_validos=num_dados_validos/num_dados_totais 
 
    return status,sem_chuva,com_chuva,sem_dados,soma_da_chuva,num_dados_totais,num_dias,num_dados_validos,per_semchuva,per_comchuva,per_semdados,per_dados_validos 
    

#
#
#
#
#
# Função para processar um arquivo CSV
#
#
#
#

def processar_arquivo(arquivo,data1="",data2=""):
    try:
        df = pd.read_csv(arquivo, parse_dates=['data'],skiprows=5)
       
        
        # Verificar se o arquivo possui a coluna "chuva"
        if 'chuva' not in df.columns:
            print(f"O arquivo '{arquivo}' não possui a coluna 'chuva' e não será processado.")
            return None
        
        # Classificar os dados em "chuva", "semchuva" e "outros=semd dados"
        df['tipo'] = df['chuva'].apply(lambda x: 'chuva' if x > 0 else ('semchuva' if x == 0 else 'outros'))

        #
        #
        # avalia se os parametros data1 e data2 foram inseridos na funcao. 
        # caso sejam nulos ele faz o tratamento de todo o peridoo do arquivo
        # caos sejam inseridos os parametros data1 e data2 calculo no periodo
        #       
        if data1 == "":
           data_inicial = df['data'].min()
        else:
           data_inicial = data1
        if data2 == "":
           data_final = df['data'].max()
        else:
           data_final = data2
           
        #
        # status=0 (tem dados)
        #       =1 (completamente sem dados)  
        
        status,sem_chuva,com_chuva,sem_dados,soma_da_chuva,num_dados_totais,num_dias,dados_validos,per_semchuva,per_comchuva,per_semdados,per_dados_validos = calcular_varios_periodo(df,data_inicial,data_final)
        
        if status == 0:  
         
            # # Calcular o máximo valor de chuva
            # # media_chuva = df[df['tipo'] == 'chuva']['chuva'].mean()
            soma_chuva = calcular_soma_periodo(df, data_inicial,data_final)
            media_chuva = calcular_media_periodo(df, data_inicial,data_final)
            desvpad_chuva = calcular_desvpad_periodo(df, data_inicial,data_final)
            variancia_chuva = calcular_variancia_periodo(df, data_inicial,data_final)
            max_chuva = calcular_variancia_periodo(df, data_inicial,data_final)
            contas_chuva = calcular_conta_periodo(df, data_inicial,data_final)
            tempo_retorno = 0.0 ##calcular_tempo_retorno(df,max_chuva)
            percentis_chuva=calcular_percentil_periodo(df,data_inicial,data_final)
        else:
            return None
        return data_inicial, data_final , num_dias,num_dados_totais,sem_chuva,com_chuva,dados_validos,sem_dados, per_semchuva,per_comchuva,per_dados_validos,per_semdados,soma_chuva,media_chuva,max_chuva,desvpad_chuva,variancia_chuva,tempo_retorno, percentis_chuva
    
    except Exception as e:
        print(f"Erro ao processar o arquivo '{arquivo}': {e}")
        return None



#-----------------------------------------------------------------------------------
#
#             PRINCIPAL 
#
#
#
#
#
#
#
#     DEFINICAO DO PERIODO CLIMA
#
#
#
inicio='1991-01-01' 
fim='2020-12-31' 
inicio_hist=datetime.strptime("3000-01-01","%Y-%m-%d")
final_hist=datetime.strptime("1900-01-01","%Y-%m-%d")

#
#
# Criar um arquivo Excel e uma planilha
#
#
#
wb = Workbook()
ws = wb.active
ws = wb.create_sheet(title="TODOS")
ws2 = wb.create_sheet(title="CLIMA")

header=["Arquivo", "Nome","Codigo","Longitude","Latitude","Altitude",
        "Data Inicial", "Data Final", "Dados Totais", "Numero de Dias", "Sem Chuva", "Com Chuva", 
        "Dados Validos","Sem Dados", "Per. Sem Chuva","Per. Com Chuva","Per. Dados Validos","Per. Sem Dados","Soma","Media",
        "Maximo Chuva", "Desvio Padrao","Variancia",
        "P05", "P10", "P20", "P25", "P50", "P75", "P90", "P95", "P99"]
ws.append(header)

ws2.append(header)


# Processar cada arquivo CSV
for arquivo in arquivos_csv:
    with open(arquivo) as file:
         station_info = [next(file).strip().split()[1] for _ in range(5)]

    # Cria um dicionário com as informações de estação
    codigo=station_info[0]
    nome=station_info[1]
    latitude=float(station_info[2]) 
    longitude=float(station_info[3]) 
    if station_info[4] == 'None':
       altitude=-999 
    else:
       altitude=float((station_info[4]))
    
    data1=""
    data2="" 
    resultado = processar_arquivo(arquivo,data1,data2)
    if resultado:
        data_inicial, data_final , num_dias,num_dados_totais,sem_chuva,com_chuva,dados_validos,sem_dados, per_semchuva,per_comchuva,per_dados_validos,per_semdados,soma_chuva,media_chuva,max_chuva,desvpad_chuva,variancia_chuva,tempo_retorno, percentis_chuva = resultado
        ws.append([arquivo, nome,codigo,longitude,latitude,altitude,data_inicial, data_final, num_dados_totais, num_dias, sem_chuva, com_chuva,dados_validos, sem_dados,
        per_semchuva,per_comchuva,per_dados_validos,per_semdados,
        soma_chuva,media_chuva,max_chuva,desvpad_chuva,variancia_chuva,tempo_retorno,
        percentis_chuva[0.05], percentis_chuva[0.10], percentis_chuva[0.20], percentis_chuva[0.25], percentis_chuva[0.50], percentis_chuva[0.75], percentis_chuva[0.90], percentis_chuva[0.95], percentis_chuva[0.99]]) 
        print(arquivo," foi processado")        
    else:
        print(arquivo," Não foi processado")

         
    # Converter as strings de data em objetos datetime
    data1 = datetime.strptime(inicio, "%Y-%m-%d")
    data2 = datetime.strptime(fim, "%Y-%m-%d")
    if ( data_inicial < inicio_hist ):
        inicio_hist=data_inicial
    
    if ( data_final > final_hist ):
        final_hist=data_final   
    
    data3_str = data1.strftime("%Y%m%d")
    data4_str = data2.strftime("%Y%m%d")
    
    data2_str = final_hist.strftime("%Y%m%d")
    data1_str = inicio_hist.strftime("%Y%m%d")
    

    # Crie o nome do arquivo de saída com base nas datas formatadas
    nome_arquivo_saida = f"relatorio_ana_{data1_str}_a_{data2_str}_e_{data3_str}_a_{data4_str}.xlsx"
    
    
    resultado = processar_arquivo(arquivo,data1,data2)
    if resultado:
        data_inicial, data_final , num_dias,num_dados_totais,sem_chuva,com_chuva,dados_validos,sem_dados, per_semchuva,per_comchuva,per_dados_validos,per_semdados,soma_chuva,media_chuva,max_chuva,desvpad_chuva,variancia_chuva,tempo_retorno, percentis_chuva = resultado
        ws2.append([arquivo, nome,codigo,longitude,latitude,altitude,data_inicial, data_final, num_dados_totais, num_dias, sem_chuva, com_chuva,dados_validos, sem_dados,
        per_semchuva,per_comchuva,per_dados_validos,per_semdados,
        soma_chuva,media_chuva,max_chuva,desvpad_chuva,variancia_chuva,tempo_retorno,
        percentis_chuva[0.05], percentis_chuva[0.10], percentis_chuva[0.20], percentis_chuva[0.25], percentis_chuva[0.50], percentis_chuva[0.75], percentis_chuva[0.90], percentis_chuva[0.95], percentis_chuva[0.99]]) 
        print(arquivo," foi processado") 
    else:
        print(arquivo," Não foi processado no periodo:",data1," a ",data2)


# Salvar o arquivo Excel com o nome gerado
wb.save(nome_arquivo_saida)
print(f"Relatório gerado e salvo como '{nome_arquivo_saida}'")
